import os
from html import escape

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DATA_FOLDER = "data"

RULES_WORKBOOK_PATH = os.path.join(
    DATA_FOLDER,
    "rules_workbook.xlsx"
)

RULES_SHEET_NAME = "Master Plan - Football Season"


def ensure_data_folder():

    os.makedirs(
        DATA_FOLDER,
        exist_ok=True
    )


def save_rules_workbook_from_upload(upload_event):

    ensure_data_folder()

    try:

        content = upload_event.content.read()

        with open(
            RULES_WORKBOOK_PATH,
            "wb"
        ) as file:

            file.write(
                content
            )

        return True

    except Exception:

        return False


def rules_workbook_exists():

    return os.path.exists(
        RULES_WORKBOOK_PATH
    )


def get_rules_workbook_path():

    return RULES_WORKBOOK_PATH


def get_cell_background_color(cell):

    try:

        fill = cell.fill

        if not fill:
            return None

        color = fill.fgColor

        if not color:
            return None

        rgb = color.rgb

        if not rgb:
            return None

        if len(rgb) == 8:

            return f"#{rgb[2:]}"

        if len(rgb) == 6:

            return f"#{rgb}"

        return None

    except Exception:

        return None


def get_cell_font_color(cell):

    try:

        color = cell.font.color

        if not color:
            return "#ffffff"

        rgb = color.rgb

        if not rgb:
            return "#ffffff"

        if len(rgb) == 8:

            return f"#{rgb[2:]}"

        if len(rgb) == 6:

            return f"#{rgb}"

        return "#ffffff"

    except Exception:

        return "#ffffff"


def get_cell_font_weight(cell):

    try:

        if cell.font.bold:

            return "bold"

        return "normal"

    except Exception:

        return "normal"


def get_cell_text_align(cell):

    try:

        alignment = cell.alignment.horizontal

        if alignment in [
            "center",
            "right",
            "left"
        ]:

            return alignment

        return "left"

    except Exception:

        return "left"


def get_cell_vertical_align(cell):

    try:

        alignment = cell.alignment.vertical

        if alignment in [
            "top",
            "center",
            "bottom"
        ]:

            return alignment

        return "middle"

    except Exception:

        return "middle"


def get_cell_value(cell):

    if cell.value is None:

        return ""

    return str(
        cell.value
    )


def get_merged_cell_map(sheet):

    merged_map = {}

    skip_cells = set()

    for merged_range in sheet.merged_cells.ranges:

        min_col = merged_range.min_col
        min_row = merged_range.min_row
        max_col = merged_range.max_col
        max_row = merged_range.max_row

        rowspan = (
            max_row
            - min_row
            + 1
        )

        colspan = (
            max_col
            - min_col
            + 1
        )

        merged_map[
            (
                min_row,
                min_col
            )
        ] = {
            "rowspan": rowspan,
            "colspan": colspan
        }

        for row in range(
            min_row,
            max_row + 1
        ):

            for col in range(
                min_col,
                max_col + 1
            ):

                if (
                    row,
                    col
                ) != (
                    min_row,
                    min_col
                ):

                    skip_cells.add(
                        (
                            row,
                            col
                        )
                    )

    return merged_map, skip_cells


def get_column_width(sheet, column_index):

    try:

        column_letter = get_column_letter(
            column_index
        )

        width = (
            sheet.column_dimensions[
                column_letter
            ].width
        )

        if not width:

            return 110

        return max(
            int(width * 8),
            70
        )

    except Exception:

        return 110


def sheet_to_html_table():

    if not rules_workbook_exists():

        return """
        <div style="
            color: white;
            background-color: #151515;
            border: 1px solid #333333;
            border-radius: 14px;
            padding: 18px;
        ">
            No rules workbook has been uploaded yet.
        </div>
        """

    try:

        workbook = load_workbook(
            RULES_WORKBOOK_PATH,
            data_only=True
        )

        if RULES_SHEET_NAME not in workbook.sheetnames:

            return f"""
            <div style="
                color: white;
                background-color: #151515;
                border: 1px solid #333333;
                border-radius: 14px;
                padding: 18px;
            ">
                Could not find sheet named: {RULES_SHEET_NAME}
            </div>
            """

        sheet = workbook[
            RULES_SHEET_NAME
        ]

        merged_map, skip_cells = get_merged_cell_map(
            sheet
        )

        max_row = sheet.max_row
        max_col = sheet.max_column

        html = """
        <div style="
            overflow-x: auto;
            width: 100%;
            background-color: #050505;
            padding-bottom: 20px;
        ">
        <table style="
            border-collapse: collapse;
            color: white;
            background-color: #050505;
            font-family: Arial, Helvetica, sans-serif;
            font-size: 14px;
            min-width: 1200px;
        ">
        """

        for row_index in range(
            1,
            max_row + 1
        ):

            row_has_content = False

            for col_index in range(
                1,
                max_col + 1
            ):

                value = sheet.cell(
                    row=row_index,
                    column=col_index
                ).value

                if value is not None and str(
                    value
                ).strip() != "":

                    row_has_content = True
                    break

            if not row_has_content:

                html += """
                <tr>
                    <td colspan="20" style="
                        height: 14px;
                        border: none;
                        background-color: #050505;
                    "></td>
                </tr>
                """

                continue

            html += "<tr>"

            for col_index in range(
                1,
                max_col + 1
            ):

                if (
                    row_index,
                    col_index
                ) in skip_cells:

                    continue

                cell = sheet.cell(
                    row=row_index,
                    column=col_index
                )

                value = get_cell_value(
                    cell
                )

                merged_info = merged_map.get(
                    (
                        row_index,
                        col_index
                    ),
                    {
                        "rowspan": 1,
                        "colspan": 1
                    }
                )

                rowspan = merged_info[
                    "rowspan"
                ]

                colspan = merged_info[
                    "colspan"
                ]

                background_color = get_cell_background_color(
                    cell
                ) or "#111111"

                font_color = get_cell_font_color(
                    cell
                )

                font_weight = get_cell_font_weight(
                    cell
                )

                text_align = get_cell_text_align(
                    cell
                )

                vertical_align = get_cell_vertical_align(
                    cell
                )

                width = get_column_width(
                    sheet,
                    col_index
                )

                escaped_value = escape(
                    value
                ).replace(
                    "\n",
                    "<br>"
                )

                if value == "":

                    escaped_value = "&nbsp;"

                html += f"""
                <td rowspan="{rowspan}" colspan="{colspan}" style="
                    border: 1px solid #444444;
                    padding: 8px;
                    min-width: {width}px;
                    max-width: 260px;
                    background-color: {background_color};
                    color: {font_color};
                    font-weight: {font_weight};
                    text-align: {text_align};
                    vertical-align: {vertical_align};
                    white-space: normal;
                    word-wrap: break-word;
                ">
                    {escaped_value}
                </td>
                """

            html += "</tr>"

        html += """
        </table>
        </div>
        """

        workbook.close()

        return html

    except Exception as error:

        return f"""
        <div style="
            color: white;
            background-color: #151515;
            border: 1px solid #ef4444;
            border-radius: 14px;
            padding: 18px;
        ">
            Error loading rules workbook: {escape(str(error))}
        </div>
        """