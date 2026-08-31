"""Excel migration/export helpers for DothPick.

PostgreSQL is the application's live source of truth when ``DATABASE_URL`` is
set. The bundled ``database.xlsx`` file is used only to seed a brand-new
database and can also be used as a human-readable export/backup. It is never
loaded over an already-populated production database.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select, text


BASE_DIR = Path(__file__).resolve().parent
BUNDLED_EXCEL_PATH = BASE_DIR / "database.xlsx"

TABLE_ORDER = [
    "users",
    "teams",
    "weeks",
    "games",
    "picks",
    "leaderboard",
    "weekly_winners",
    "settings",
    "system_logs",
]


def _metadata():
    # Imported lazily to avoid a database.db <-> database.models import cycle.
    from database.models import Base

    return Base.metadata


def _is_real_workbook(path: Path) -> bool:
    if not path.exists() or path.stat().st_size < 4:
        return False

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        has_sheets = bool(workbook.sheetnames)
        workbook.close()
        return has_sheets
    except Exception:
        return False


def _convert_for_column(value, column):
    if value is None:
        return None

    python_type = None
    try:
        python_type = column.type.python_type
    except Exception:
        pass

    if python_type is bool:
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "y"}
        return bool(value)

    if python_type is int:
        return int(value)

    if python_type is datetime:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        if isinstance(value, str) and value.strip():
            try:
                return datetime.fromisoformat(value.strip())
            except ValueError:
                return None

    if python_type is str and not isinstance(value, str):
        return str(value)

    return value


def _reset_postgres_id_sequences(connection, metadata) -> None:
    """Advance SERIAL sequences after importing explicit primary-key IDs."""

    if connection.dialect.name != "postgresql":
        return

    for table_name in TABLE_ORDER:
        table = metadata.tables.get(table_name)
        if table is None or "id" not in table.c:
            continue

        # Table names are internal constants from SQLAlchemy metadata, not
        # user input. Quote them through the dialect before building SQL.
        quoted_table = connection.dialect.identifier_preparer.quote(table_name)

        connection.execute(
            text(
                f"""
                SELECT setval(
                    pg_get_serial_sequence(:table_name, 'id'),
                    GREATEST(COALESCE(MAX(id), 0), 1),
                    COALESCE(MAX(id), 0) > 0
                )
                FROM {quoted_table}
                """
            ),
            {"table_name": table_name},
        )


def import_xlsx_into_database(
    engine,
    xlsx_path: Path | str = BUNDLED_EXCEL_PATH,
    *,
    replace_existing: bool = False,
) -> bool:
    """Import the workbook into the configured SQL database.

    ``replace_existing`` defaults to False for safety. Startup code calls this
    function only when the users table is empty, so a Render redeploy can never
    overwrite live Neon picks with the bundled Excel snapshot.

    Returns True if at least one data row was imported.
    """

    xlsx_path = Path(xlsx_path)

    if not _is_real_workbook(xlsx_path):
        return False

    metadata = _metadata()
    workbook = load_workbook(xlsx_path, data_only=False)
    known_sheets = [name for name in TABLE_ORDER if name in workbook.sheetnames]

    if not known_sheets:
        workbook.close()
        return False

    loaded_any_rows = False

    try:
        with engine.begin() as connection:
            if replace_existing:
                # Child tables first so foreign-key constraints remain valid.
                for table_name in reversed(TABLE_ORDER):
                    table = metadata.tables.get(table_name)
                    if table is not None:
                        connection.execute(delete(table))

            for table_name in TABLE_ORDER:
                if table_name not in workbook.sheetnames:
                    continue

                table = metadata.tables.get(table_name)
                if table is None:
                    continue

                # When this is a safe, non-destructive seed, do not add rows to
                # a table that already contains data.
                if not replace_existing:
                    existing = connection.execute(
                        select(table.c.id).limit(1)
                    ).first()
                    if existing is not None:
                        continue

                sheet = workbook[table_name]
                rows = sheet.iter_rows(values_only=True)
                header_row = next(rows, None)
                if not header_row:
                    continue

                headers = [
                    str(value).strip() if value is not None else ""
                    for value in header_row
                ]
                valid_headers = {column.name for column in table.columns}

                records = []
                for row in rows:
                    if not any(value is not None for value in row):
                        continue

                    record = {}
                    for index, header in enumerate(headers):
                        if (
                            not header
                            or header not in valid_headers
                            or index >= len(row)
                        ):
                            continue

                        column = table.columns[header]
                        record[header] = _convert_for_column(row[index], column)

                    if record:
                        records.append(record)

                if records:
                    connection.execute(table.insert(), records)
                    loaded_any_rows = True

            _reset_postgres_id_sequences(connection, metadata)

        return loaded_any_rows

    finally:
        workbook.close()


def export_database_to_xlsx(
    engine,
    output_path: Path | str = BUNDLED_EXCEL_PATH,
) -> Path:
    """Export all application tables to an Excel workbook.

    This is an explicit export/backup helper. It is intentionally not called
    after every commit because PostgreSQL/Neon is the live database.
    """

    output_path = Path(output_path)
    metadata = _metadata()

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    with engine.connect() as connection:
        for table_name in TABLE_ORDER:
            table = metadata.tables.get(table_name)
            if table is None:
                continue

            sheet = workbook.create_sheet(title=table_name)
            columns = list(table.columns)
            sheet.append([column.name for column in columns])

            statement = select(table)
            if "id" in table.c:
                statement = statement.order_by(table.c.id)

            for row in connection.execute(statement).mappings():
                values = []
                for column in columns:
                    value = row[column.name]
                    if isinstance(value, datetime) and value.tzinfo is not None:
                        value = value.replace(tzinfo=None)
                    values.append(value)

                sheet.append(values)

            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()

    return output_path
